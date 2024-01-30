import abc
import csv
import json
import locale
import logging
import os.path
from datetime import datetime
from enum import Enum
from json import JSONDecodeError
from pathlib import Path

import dateparser
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from reader.etf_reader_configs import EtfReaderConfigs
from report.region import Region

try:
    locale.setlocale(locale.LC_ALL, 'de_DE.utf8')
except locale.Error as e:
    print("Could not load local 'de_DE.utf8'")

log = logging.getLogger("ac")


class FundFamily(Enum):
    ISHARES = 'ISHARES'
    LGIM = "LGIM"
    SPDR = "SPDR"
    VANECK = 'VANECK'
    VANGUARD = 'VANGUARD'
    XTRACKERS = "XTRACKERS"


class LocationCodes(Enum):
    DE_FULL_NAME = "de_full_name.json"
    EN_FULL_NAME = "en_full_name.json"
    ALPHA_2_CODE = "alpha_2_code.json"


class FileTypes(Enum):
    XLSX = "xlsx"
    CSV = "csv"


class EtfReader:
    NAME_COUNTER = 0
    ISIN_COUNTER = 0
    ISIN_LOOKUP = None
    ISIN_TO_NAME_LOOKUP = None
    NOT_EXIST = '----'
    REGION_MAPPING = {}

    def __init__(self, fpath: str, config_name: str = None):
        self.name = EtfReader.NOT_EXIST
        self.fpath = fpath
        self.asset = None
        self.sheet = None
        self.values = []
        self.isin = ''
        self.isin_src = None
        self.fund_family = None
        self.config_name = config_name
        self.date_format = None
        self.date_parse = None
        self.date_src = None
        self.file_type = None
        self.raw_data = None

        if "xlsx" in self.fpath:
            self.file_type = FileTypes.XLSX
        else:
            self.file_type = FileTypes.CSV

    def init_from_config(self):
        self.find_config()
        config = EtfReaderConfigs.get_config(self.fund_family, self.config_name)

        inc = 0
        if "xlsx" in config["BASE"]["regex"]:
            inc = 1

        self.start_row = int(config["LINES"]["start_row"]) + inc
        self.ticker_col = int(config["LINES"]["ticker_col"]) + inc
        self.holding_name_col = int(config["LINES"]["name_col"]) + inc
        self.weight_col = int(config["LINES"]["weight_col"]) + inc
        self.region_col = int(config["LINES"]["region_col"]) + inc

        if config.has_option("DATE", "src"):
            self.date_src = config.get("DATE", "src")
        else:
            self.date_row = int(config["DATE"]["row"]) + inc
            self.date_col = int(config["DATE"]["col"]) + inc

        self.covert_str = config["BASE"]["number_convert"]
        self.location_code = config["BASE"]["location_code"]

        if config.has_option("DATE", "format"):
            self.date_format = config["DATE"]["format"]
        if config.has_option("DATE", "parse"):
            self.date_parse = config.get("DATE", "parse", raw=True)

        if config.has_section("ETF_NAME"):
            self.name_row = int(config["ETF_NAME"]["row"]) + inc
            self.name_col = int(config["ETF_NAME"]["col"]) + inc
        else:
            self.name_row = self.name_col = None

        if config.has_option("ETF_ISIN", "row"):
            self.isin_row = int(config["ETF_ISIN"]["row"]) + inc
            self.isin_col = int(config["ETF_ISIN"]["col"]) + inc
        elif config.has_option("ETF_ISIN", "src"):
            self.isin_src = config.get("ETF_ISIN", "src")
        else:
            self.isin_row = self.isin_col = None

    def find_config(self) -> bool:
        if self.config_name:
            return True

        for cfg_name, config in EtfReaderConfigs.get_family_configs(self.fund_family).items():
            line = int(config["SUB_DETECTION"]["row"])
            col = int(config["SUB_DETECTION"]["col"])
            val = config["SUB_DETECTION"]["value"]

            if val == self.get_data(line, col):
                log.debug(f"... => {cfg_name}")
                self.config_name = cfg_name
                return True
        return False

    def convert_str_to_float(self, nbr_str: str):
        if self.covert_str == "COMMA":
            nbr_str = nbr_str.replace(",", ".")
        nbr_str = nbr_str.replace("%", "")
        return float(nbr_str.strip())

    def get_name(self):
        if self.name_row:
            return self.get_data(self.name_row, self.name_col)

        name = self.get_name_from_isin(self.fund_family, self.isin)

        if name != EtfReader.NOT_EXIST:
            return name

        return self.set_default_name()

    def get_isin(self):
        if self.isin_src:
            if self.isin_src == "FILE_NAME":
                return os.path.basename(self.fpath)
        return self.get_data(self.isin_row, self.isin_col)

    def get_date(self):
        if not self.date_src:
            cell_value = self.get_data(self.date_row, self.date_col)
        elif self.date_src == "SHEET_TITLE":
            cell_value = self.sheet.title

        if self.date_format:
            cell_value = eval(f"'{cell_value}'{self.date_format}")

        if self.date_parse:
            return datetime.strptime(cell_value, self.date_parse)

        date_obj = self.parse_date(cell_value)
        return date_obj

    def update_region(self, region: str, weight: float) -> None:
        if region not in self.asset.regions:
            self.asset.regions[region] = Region(region, weight, 1)
        else:
            self.asset.regions[region].weight += weight
            self.asset.regions[region].num_of_countries += 1

    def open_file(self):
        if self.file_type == FileTypes.XLSX:
            self.read_sheet_from_wb()
        else:
            self.read_csv()

    def read_sheet_from_wb(self):
        try:
            wb = load_workbook(filename=self.fpath)
            self.sheet = wb.active
            self.raw_data = wb.active
        except InvalidFileException as e:
            log.error(f"Could not read file: {self.fpath}")

    def read_csv(self):
        with open(self.fpath, newline='', encoding="utf-8-sig") as csvfile:
            self.raw_data = list(csv.reader(csvfile, delimiter=',', quotechar='"'))

    def get_row_count(self):
        if self.file_type == FileTypes.XLSX:
            return self.raw_data.max_row
        else:
            return len(self.raw_data)

    def get_data(self, row: int, col: int):
        try:
            if self.file_type == FileTypes.XLSX:
                return self.raw_data.cell(row, col).value
            else:
                return self.raw_data[row][col]
        except IndexError:
            return None

    def parse_date(self, last_update: str, date_formats: list[str] = None) -> datetime:
        try:
            date_obj = dateparser.parse(last_update, date_formats=date_formats)
            return date_obj if date_obj else datetime.now()
        except ValueError:
            return datetime.now()

    @staticmethod
    def get_isin_from_file_name(fund_family, name) -> str:
        if EtfReader.ISIN_LOOKUP is None:
            path_to_mapping = Path(__file__).parent
            path_to_mapping = os.path.join(path_to_mapping, "../", "mappings", "lookups", "isin_lookup.json")
            EtfReader.ISIN_LOOKUP = EtfReader.read_json(path_to_mapping)
        return EtfReader.ISIN_LOOKUP.get(fund_family.value, {}).get(name, EtfReader.NOT_EXIST)

    @staticmethod
    def get_name_from_isin(fund_family, isin) -> str:
        if EtfReader.ISIN_TO_NAME_LOOKUP is None:
            path_to_mapping = Path(__file__).parent
            path_to_mapping = os.path.join(path_to_mapping, "../", "mappings", "lookups", "isin_to_name_lookup.json")
            EtfReader.ISIN_TO_NAME_LOOKUP = EtfReader.read_json(path_to_mapping)
        return EtfReader.ISIN_TO_NAME_LOOKUP.get(fund_family.value, {}).get(isin, EtfReader.NOT_EXIST)

    @staticmethod
    def get_isin_and_names():
        # pre load lookup files
        EtfReader.get_name_from_isin(FundFamily.VANGUARD, '')
        EtfReader.get_isin_from_file_name(FundFamily.VANGUARD, '')

        etf_data = EtfReader.ISIN_TO_NAME_LOOKUP
        for issuer, data in EtfReader.ISIN_LOOKUP.items():
            if issuer in etf_data:
                continue

            etf_data[issuer] = {}
            for name, isin in data.items():
                etf_data[issuer][isin] = name

        return etf_data

    @staticmethod
    def read_json(path: str):
        if not os.path.exists(path):
            log.warning(f"File not exists:{path}")
            return {}

        with open(path, 'r', encoding="utf-8") as f:
            try:
                return json.load(f)
            except JSONDecodeError:
                log.error(f"Could not read region mapping for: {path}")
                return {}

    @staticmethod
    def get_region_code(location_code, name):
        if location_code not in EtfReader.REGION_MAPPING:
            path_to_mapping = Path(__file__).parent
            path_to_mapping = os.path.join(path_to_mapping, "../", "mappings", "location_codes", location_code.value)
            EtfReader.REGION_MAPPING[location_code] = EtfReader.read_json(path_to_mapping)
        return EtfReader.REGION_MAPPING[location_code].get(name, name)

    def set_default_isin(self):
        EtfReader.ISIN_COUNTER += 1
        self.isin = f'XX{EtfReader.ISIN_COUNTER:09d}0'
        self.asset.isin = self.isin

    def set_default_name(self):
        EtfReader.NAME_COUNTER += 1
        self.name = f'{self.fund_family.value} ETF{EtfReader.NAME_COUNTER:02d}'
        return self.name
        # self.asset.name = self.name

    @abc.abstractmethod
    def read_sheet(self):
        """Method documentation"""
        return

    @abc.abstractmethod
    def read_asset(self):
        """Method documentation"""
        return
